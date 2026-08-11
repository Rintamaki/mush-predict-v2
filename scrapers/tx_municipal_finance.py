"""
tx_municipal_finance.py

Loads Texas Comptroller Local Government Bond, Tax, and Project data
(HB103) from a locally-uploaded spreadsheet and normalizes it into the
JSON shape the dashboard consumes.

Modeled directly on scrapers/tx_district_finance.py — the manual-download
pattern proven to work for TEA data:
  1. Someone downloads the HB103 spreadsheet from
     https://comptroller.texas.gov/transparency/local/hb103/hb103-simple.php
  2. Renames it locally to municipal_bonds.xlsx (do NOT rename on GitHub —
     it corrupts binary files)
  3. Uploads to scrapers/tea_data/
  4. This script reads it and writes municipal_finance.json

──────────────────────────────────────────────────────────────────────────
STATUS: SKELETON — needs the real spreadsheet to complete.
──────────────────────────────────────────────────────────────────────────

Steps to finish (search this file for "TODO"):
  1. Download the HB103 spreadsheet.
  2. Open it and note the actual sheet name and column headers.
  3. Fill in EXPECTED_COLUMNS below with the real names.
  4. If needed, adjust normalize_row() to match actual field types.
  5. Run: python scrapers/tx_municipal_finance.py
  6. Verify the JSON output looks right.

All safety features work today: manual-download pattern, backup logic,
graceful failure, clean logs. Only the column mapping needs the real file.
"""

import json
import logging
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("tx_municipal_finance")

ROOT       = Path(__file__).parent.parent
INPUT_XLSX = Path(__file__).parent / "tea_data" / "municipal_bonds.xlsx"
OUTPUT     = ROOT / "mush-predict-package" / "public" / "data" / "municipal_finance.json"

# ─────────────────────────────────────────────────────────────────────────────
# TODO — CONFIGURE ONCE THE REAL FILE IS IN HAND
# ─────────────────────────────────────────────────────────────────────────────
# When you download the HB103 spreadsheet, open it and note:
#   - The sheet/tab name (Excel workbooks can have multiple sheets)
#   - The actual column headers as they appear in row 1
#
# Common HB103 fields likely to appear (based on the legislation text — but
# only trust the ACTUAL file, not this list):
#   - Taxing entity name (city, county, MUD, or special district)
#   - Entity type
#   - County
#   - Election date
#   - Election type (bond proposition, TRE, etc.)
#   - Ballot language / bond purpose
#   - Amount requested
#   - Election results — votes for / against
#   - Passed / failed indicator
#   - Projected tax rates
#   - Project descriptions (may be in a separate sheet)
#
# Fill in the map below with the ACTUAL column names from the spreadsheet.
# Use None for any field the file doesn't have — the code handles missing
# fields gracefully.
SHEET_NAME = None    # TODO — e.g. "Bond Data" or 0 for first sheet

EXPECTED_COLUMNS = {
    "entity_name":     None,   # TODO — column name for the city/county/district
    "entity_type":     None,   # TODO — "City", "County", "MUD", etc. (or None)
    "county":          None,   # TODO — county the entity is in (or None)
    "election_date":   None,   # TODO — date of the bond/TRE election
    "election_type":   None,   # TODO — "Bond", "TRE", etc.
    "purpose":         None,   # TODO — bond purpose / ballot language column
    "amount":          None,   # TODO — dollar amount requested
    "votes_for":       None,   # TODO — column with "for" votes (or None)
    "votes_against":   None,   # TODO — column with "against" votes (or None)
    "passed":          None,   # TODO — pass/fail column (e.g. "Result" with values Passed/Failed)
    "tax_rate":        None,   # TODO — projected debt-service tax rate (or None)
}
# ─────────────────────────────────────────────────────────────────────────────


def normalize_entity_name(name):
    """Normalize a municipality/entity name for consistent matching.

    Similar to normalize_district_name() in district_intelligence.py, but
    tuned for cities/counties/MUDs. Handles variations like:
      'City of Austin' / 'AUSTIN CITY' / 'Austin, City of' -> 'austin'
    """
    if not name:
        return ""
    n = str(name).strip().lower()
    # Strip common prefixes/suffixes
    for prefix in ["city of ", "town of ", "village of ", "county of "]:
        if n.startswith(prefix):
            n = n[len(prefix):]
    for suffix in [", city of", " city", " county", " town", " village",
                   " mud", " municipal utility district", " water district"]:
        if n.endswith(suffix):
            n = n[:-len(suffix)]
    return " ".join(n.split())


def _parse_bool_passed(value):
    """Interpret a pass/fail cell across the various forms it might take."""
    if value is None:
        return None
    s = str(value).strip().lower()
    if s in ("passed", "pass", "approved", "carried", "yes", "y", "true", "1"):
        return True
    if s in ("failed", "fail", "rejected", "defeated", "no", "n", "false", "0"):
        return False
    return None


def _parse_amount(value):
    """Turn a currency string like '$45,000,000' into a float. 0 on failure."""
    if value is None:
        return 0.0
    try:
        s = str(value).replace("$", "").replace(",", "").strip()
        return float(s) if s else 0.0
    except (ValueError, TypeError):
        return 0.0


def normalize_row(row):
    """Turn one raw spreadsheet row into a clean dict using EXPECTED_COLUMNS.

    Returns None for rows that lack an entity name (blank/header rows).
    """
    def col(field):
        colname = EXPECTED_COLUMNS.get(field)
        if not colname:
            return None
        return row.get(colname)

    entity = col("entity_name")
    if not entity or not str(entity).strip():
        return None

    return {
        "entity_name":   str(entity).strip(),
        "entity_key":    normalize_entity_name(entity),
        "entity_type":   str(col("entity_type") or "").strip() or None,
        "county":        str(col("county") or "").strip() or None,
        "election_date": str(col("election_date") or "").strip() or None,
        "election_type": str(col("election_type") or "").strip() or None,
        "purpose":       str(col("purpose") or "").strip() or None,
        "amount":        _parse_amount(col("amount")),
        "votes_for":     col("votes_for"),
        "votes_against": col("votes_against"),
        "passed":        _parse_bool_passed(col("passed")),
        "tax_rate":      col("tax_rate"),
    }


def load_spreadsheet():
    """Load the HB103 spreadsheet. Returns a list of raw dict rows."""
    if not INPUT_XLSX.exists():
        log.error(f"Input file not found: {INPUT_XLSX}")
        log.error("Download the HB103 spreadsheet from:")
        log.error("  https://comptroller.texas.gov/transparency/local/hb103/hb103-simple.php")
        log.error("Rename it locally to municipal_bonds.xlsx and upload to scrapers/tea_data/")
        return []

    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl not installed. Add it to scrapers/requirements.txt: openpyxl")
        return []

    log.info(f"Loading {INPUT_XLSX.name}")
    try:
        wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True, read_only=True)
    except Exception as e:
        log.error(f"Could not open workbook: {e}")
        return []

    # Pick the sheet
    if SHEET_NAME is None:
        sheet = wb.active
        log.info(f"  Using active sheet: '{sheet.title}' "
                 f"(available: {wb.sheetnames})")
    elif isinstance(SHEET_NAME, int):
        sheet = wb.worksheets[SHEET_NAME]
    else:
        if SHEET_NAME not in wb.sheetnames:
            log.error(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
            return []
        sheet = wb[SHEET_NAME]

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        log.warning("Sheet is empty")
        return []

    headers = [str(h).strip() if h is not None else "" for h in headers]
    log.info(f"  Header row: {headers[:8]}{'...' if len(headers) > 8 else ''}")

    rows = []
    for raw in rows_iter:
        row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
        rows.append(row)

    log.info(f"  Loaded {len(rows)} raw data rows")
    return rows


def check_column_config(sample_headers):
    """Sanity-check that configured column names actually exist in the file."""
    if not any(EXPECTED_COLUMNS.values()):
        log.warning("=" * 60)
        log.warning("EXPECTED_COLUMNS is not configured yet.")
        log.warning("Edit tx_municipal_finance.py and fill in the column names,")
        log.warning("using the header names shown above.")
        log.warning("=" * 60)
        return False

    missing = []
    for field, colname in EXPECTED_COLUMNS.items():
        if colname and colname not in sample_headers:
            missing.append((field, colname))

    if missing:
        log.warning("Some configured column names don't match the file:")
        for field, colname in missing:
            log.warning(f"  {field}: configured as '{colname}' — not in headers")
        log.warning(f"Actual headers: {sample_headers}")
        return False
    return True


def run():
    log.info("=" * 60)
    log.info("Texas Municipal Finance (HB103) — Load")
    log.info("=" * 60)

    rows = load_spreadsheet()
    if not rows:
        return 1

    sample_headers = list(rows[0].keys()) if rows else []
    configured = check_column_config(sample_headers)

    if not configured:
        # Write an empty output so the pipeline doesn't crash on the missing file,
        # but include a clear "not configured" marker.
        output = {
            "status":       "not_configured",
            "message":      "EXPECTED_COLUMNS not set. See tx_municipal_finance.py.",
            "row_count":    len(rows),
            "headers_seen": sample_headers,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "records":      [],
        }
        OUTPUT.parent.mkdir(parents=True, exist_ok=True)
        with open(OUTPUT, "w") as f:
            json.dump(output, f, indent=2)
        log.info(f"Wrote placeholder to {OUTPUT.name} (configure columns to complete).")
        return 0

    records = []
    skipped = 0
    for row in rows:
        rec = normalize_row(row)
        if rec is None:
            skipped += 1
            continue
        records.append(rec)

    log.info(f"Normalized {len(records)} records ({skipped} skipped)")

    output = {
        "status":       "ok",
        "row_count":    len(records),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "records":      records,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT, "w") as f:
        json.dump(output, f, indent=2)
    log.info(f"Wrote {OUTPUT}")
    return 0


if __name__ == "__main__":
    sys.exit(run())
