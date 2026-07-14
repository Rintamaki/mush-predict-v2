"""
tx_district_finance.py

Parses TEA (Texas Education Agency) PEIMS financial data to identify
Texas school districts by their facility-related spending.

NOTE ON DATA SOURCE:
TEA's website blocks automated downloads from cloud/datacenter IP ranges
(including GitHub Actions runners), even though the data itself is fully
public. Rather than fight that block, this script reads the xlsx from a
file committed to the repo. Since TEA only republishes this file about
once a year, the workflow is:

  1. Once a year, download the xlsx yourself in a browser from:
     https://tea.texas.gov/finance-and-grants/state-funding/state-funding-reports-and-data/peims-financial-data-downloads
  2. Upload it to: scrapers/tea_data/summarized-financial-data.xlsx
     (always use this exact filename regardless of what TEA names it)
  3. Commit it to the repo
  4. Run this script (manually or via the scheduled workflow) — it reads
     the local file and regenerates tx_district_finance.json

FUNCTION CODES / COLUMNS that matter for McKinstry work (matched directly
by column name in TEA's summarized DATAMART file):
  ALL FUNDS-PLANT MAINTENANCE/OPERA EXPEND, FCT51   = Plant M&O (HVAC, custodial, utilities)
  ALL FUNDS-SECURITY/MONITORING SERVICE EXPEND, FCT52 = Security & Monitoring
  ALL FUNDS-CAPITAL PROJECTS(OBJECT 6600) FOR TD    = Capital/bond-funded construction
    (closest proxy available — TEA's summarized file doesn't break out
    Function 81 "Facilities Acquisition & Construction" as its own column,
    so we use Capital Projects fund spending instead, which captures the
    same bond-funded construction activity)
"""

import io
import json
import logging
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s  %(levelname)-8s %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger('tea')

# ── Config ──────────────────────────────────────────────────────────────────
YEARS_TO_KEEP = 3

SCRIPT_DIR   = Path(__file__).parent
LOCAL_XLSX   = SCRIPT_DIR / 'tea_data' / 'summarized-financial-data.xlsx'
OUTPUT_FILE  = SCRIPT_DIR.parent / 'mush-predict-package' / 'public' / 'data' / 'tx_district_finance.json'


# ── Load local file ─────────────────────────────────────────────────────────
def load_local_xlsx():
    """Read the manually-uploaded xlsx from the repo."""
    if not LOCAL_XLSX.exists():
        log.error(f'File not found: {LOCAL_XLSX}')
        log.error('')
        log.error('This scraper reads a manually-downloaded file rather than fetching')
        log.error('from TEA directly (their site blocks automated/cloud IP downloads).')
        log.error('')
        log.error('To fix this:')
        log.error('  1. Go to https://tea.texas.gov/finance-and-grants/state-funding/'
                   'state-funding-reports-and-data/peims-financial-data-downloads')
        log.error('  2. Download the "Summarized PEIMS Actual Financial Data" xlsx')
        log.error(f'  3. Upload it to the repo at: scrapers/tea_data/summarized-financial-data.xlsx')
        log.error('  4. Commit and re-run this workflow')
        return None

    size_mb = LOCAL_XLSX.stat().st_size / 1e6
    log.info(f'Reading local file: {LOCAL_XLSX} ({size_mb:.1f} MB)')
    with open(LOCAL_XLSX, 'rb') as f:
        return f.read()


# ── Parse ───────────────────────────────────────────────────────────────────
def parse_tea_xlsx(data: bytes):
    try:
        import openpyxl
    except ImportError:
        log.error('openpyxl not installed — add openpyxl to requirements.txt')
        return None

    log.info('Parsing xlsx (this can take a minute for large files)…')
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    log.info(f'Opened workbook: sheet "{ws.title}", {ws.max_row} rows')

    header = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    log.info(f'Detected {len(header)} columns. Full list:')
    for i, h in enumerate(header):
        log.info(f'  [{i:3d}] {h}')

    def find_col(candidates):
        low_hdr = [h.lower() for h in header]
        for name in candidates:
            for i, h in enumerate(low_hdr):
                if name.lower() in h:
                    return i
        return None

    col_dist_id   = find_col(['district number', 'district id'])
    col_dist_name = find_col(['district name'])
    col_year      = find_col(['year'])

    # These are exact column matches based on the confirmed TEA DATAMART layout.
    # Using "ALL FUNDS" versions (not "GEN FUNDS") since All Funds includes
    # state/federal revenue sources in addition to local — a fuller picture
    # of actual facility spending.
    col_plant_maint = find_col(['all funds-plant maintenance'])
    col_security    = find_col(['all funds-security/monitoring'])
    col_capital     = find_col(['all funds-capital projects(object 6600)'])

    log.info(
        f'Column mapping: '
        f'dist_id={col_dist_id}, dist_name={col_dist_name}, year={col_year}, '
        f'plant_maint={col_plant_maint}({header[col_plant_maint] if col_plant_maint is not None else "?"}), '
        f'security={col_security}({header[col_security] if col_security is not None else "?"}), '
        f'capital={col_capital}({header[col_capital] if col_capital is not None else "?"})'
    )

    if None in (col_dist_id, col_dist_name, col_year, col_plant_maint, col_security, col_capital):
        log.error('Could not locate all required columns')
        log.error(f'Available columns: {header}')
        return None

    districts = defaultdict(lambda: {'name': '', 'years': defaultdict(lambda: defaultdict(float))})
    row_count = 0
    matched_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1

        try:
            dist_id_raw = row[col_dist_id]
            dist_name   = str(row[col_dist_name]).strip() if row[col_dist_name] else ''
            year_raw    = row[col_year]

            if dist_id_raw is None or year_raw is None:
                continue

            dist_id  = str(dist_id_raw).strip()
            year     = str(year_raw).strip()

            plant_amt    = float(row[col_plant_maint]) if row[col_plant_maint] else 0.0
            security_amt = float(row[col_security]) if row[col_security] else 0.0
            capital_amt  = float(row[col_capital]) if row[col_capital] else 0.0
        except (ValueError, TypeError, IndexError):
            continue

        if not dist_id or not year:
            continue

        matched_rows += 1
        year_clean = year[:9] if len(year) >= 9 else year

        d = districts[dist_id]
        if not d['name'] and dist_name:
            d['name'] = dist_name
        d['years'][year_clean]['plant_maintenance']       += plant_amt
        d['years'][year_clean]['security_monitoring']      += security_amt
        d['years'][year_clean]['facilities_construction']  += capital_amt

    log.info(f'Parsed {row_count:,} rows, {matched_rows:,} had valid district/year data')
    log.info(f'Aggregated into {len(districts)} unique districts')
    return dict(districts)


# ── Build output ────────────────────────────────────────────────────────────
def build_output(districts_data):
    if not districts_data:
        return None

    all_years = set()
    for d in districts_data.values():
        all_years.update(d['years'].keys())
    sorted_years = sorted(all_years, reverse=True)[:YEARS_TO_KEEP]
    log.info(f'Keeping years: {sorted_years}')

    records = []
    for dist_id, d in districts_data.items():
        if not d['name']:
            continue

        history = []
        for y in sorted(sorted_years):
            year_data = d['years'].get(y, {})
            if not year_data:
                continue
            history.append({
                'year': y,
                'plant_maintenance':       round(year_data.get('plant_maintenance', 0), 0),
                'security_monitoring':     round(year_data.get('security_monitoring', 0), 0),
                'facilities_construction': round(year_data.get('facilities_construction', 0), 0),
            })

        if not history:
            continue

        latest = history[-1]
        total_facility = (
            latest['plant_maintenance'] +
            latest['security_monitoring'] +
            latest['facilities_construction']
        )
        if total_facility < 100000:
            continue

        records.append({
            'district_id':   dist_id,
            'district_name': d['name'],
            'latest_year':   latest['year'],
            'spending':      {**latest, 'total_facility': total_facility},
            'history':       history,
        })

    records.sort(key=lambda r: r['spending']['total_facility'], reverse=True)
    for i, r in enumerate(records):
        r['rank'] = i + 1

    return {
        'generated_at':   datetime.utcnow().isoformat() + 'Z',
        'source':         'Texas Education Agency PEIMS Summarized Actual Financial Data',
        'source_url':     'https://tea.texas.gov/finance-and-grants/state-funding/state-funding-reports-and-data/peims-financial-data-downloads',
        'years_included': sorted_years,
        'notes': (
            'Annual actual expenditures per Texas ISD, from TEA\'s summarized PEIMS DATAMART. '
            'Data is 1-2 years behind current. "Plant Maintenance" = Function 51 (HVAC, custodial, '
            'utilities, ALL FUNDS). "Security" = Function 52 (ALL FUNDS). "Facilities Construction" = '
            'Capital Projects fund spending (Object 6600, ALL FUNDS) — this is the closest available '
            'proxy for bond-funded construction; TEA\'s summarized file does not break out Function 81 '
            'as a discrete column. Source file is manually downloaded and committed to the repo '
            '(TEA blocks automated fetches from cloud/datacenter IPs).'
        ),
        'district_count': len(records),
        'districts':      records,
    }


# ── Main ────────────────────────────────────────────────────────────────────
def run():
    log.info('=' * 60)
    log.info('TEA District Finance Scraper (local file mode)')
    log.info('=' * 60)

    raw = load_local_xlsx()
    if not raw:
        return 1

    districts = parse_tea_xlsx(raw)
    if not districts:
        return 1

    output = build_output(districts)
    if not output:
        return 1

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, 'w') as f:
        json.dump(output, f, indent=2)

    log.info(f'✅ Wrote {output["district_count"]} districts to {OUTPUT_FILE}')
    log.info('Top 5 by facility spend (latest year):')
    for r in output['districts'][:5]:
        log.info(f'  #{r["rank"]}  {r["district_name"]:40s}  ${r["spending"]["total_facility"]/1e6:,.1f}M')

    return 0


if __name__ == '__main__':
    sys.exit(run())
