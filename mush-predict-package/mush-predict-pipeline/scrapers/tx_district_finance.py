"""
tx_district_finance.py

Scrapes TEA (Texas Education Agency) PEIMS financial data to identify
Texas school districts by their facility-related spending.

TEA changes the xlsx URL periodically (they embed the file date in the URL).
So we first scrape the download page, find the xlsx link dynamically, then
download and parse it.

FUNCTION CODES that matter for McKinstry work:
  51 = Plant Maintenance & Operations (HVAC, custodial, utilities)
  52 = Security & Monitoring Services
  81 = Facilities Acquisition & Construction (capital projects, bonds)
"""

import io
import json
import logging
import os
import re
import sys
import urllib.request
from collections import defaultdict
from datetime import datetime
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s  %(levelname)-8s %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger('tea')

# ── Config ──────────────────────────────────────────────────────────────────
TEA_DOWNLOADS_PAGE = (
    'https://tea.texas.gov/finance-and-grants/state-funding/'
    'state-funding-reports-and-data/peims-financial-data-downloads'
)

FACILITY_FUNCTIONS = {
    '51': 'plant_maintenance',
    '52': 'security_monitoring',
    '81': 'facilities_construction',
}

YEARS_TO_KEEP = 3

OUTPUT_FILE = Path(__file__).parent.parent / 'mush-predict-package' / 'public' / 'data' / 'tx_district_finance.json'

USER_AGENT = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'


# ── Discover the current xlsx URL ───────────────────────────────────────────
def find_xlsx_url():
    """Scrape the TEA downloads page to find the summarized xlsx link."""
    # Allow env-var override so this doesn't break if the whole page moves
    if os.environ.get('TEA_XLSX_URL'):
        url = os.environ['TEA_XLSX_URL']
        log.info(f'Using TEA_XLSX_URL from env: {url}')
        return url

    log.info(f'Scraping TEA downloads page for current xlsx URL')
    req = urllib.request.Request(TEA_DOWNLOADS_PAGE, headers={'User-Agent': USER_AGENT})
    with urllib.request.urlopen(req, timeout=30) as resp:
        html = resp.read().decode('utf-8', errors='ignore')

    # Look for an href pointing at a summarized financial data xlsx.
    # Pattern: matches URLs like
    #   /reports-and-data/financial-reports/.../2009-2025-summarized-financial-data-04-08-2026.xlsx
    patterns = [
        r'href="(https://tea\.texas\.gov/[^"]*summarized[^"]*\.xlsx)"',
        r'href="(/[^"]*summarized[^"]*financial[^"]*\.xlsx)"',
        r'href="([^"]*summarized-financial-data[^"]*\.xlsx)"',
    ]
    for pattern in patterns:
        matches = re.findall(pattern, html, re.IGNORECASE)
        if matches:
            url = matches[0]
            if url.startswith('/'):
                url = 'https://tea.texas.gov' + url
            log.info(f'Found xlsx URL: {url}')
            return url

    log.error('Could not find any summarized financial xlsx link on the TEA page')
    log.error('Set TEA_XLSX_URL env var to override, or check the page manually:')
    log.error(f'  {TEA_DOWNLOADS_PAGE}')
    return None


# ── Download ────────────────────────────────────────────────────────────────
def download_tea_data(url):
    log.info(f'Downloading TEA PEIMS data from {url}')
    req = urllib.request.Request(url, headers={'User-Agent': USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            data = resp.read()
            log.info(f'Downloaded {len(data)/1e6:.1f} MB')
            return data
    except Exception as e:
        log.error(f'TEA download failed: {e}')
        return None


# ── Parse ───────────────────────────────────────────────────────────────────
def parse_tea_xlsx(data: bytes):
    try:
        import openpyxl
    except ImportError:
        log.error('openpyxl not installed — add openpyxl to requirements.txt')
        return None

    log.info('Parsing xlsx (this can take a minute)…')
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    log.info(f'Opened workbook: sheet "{ws.title}", {ws.max_row} rows')

    # Discover column layout from header row
    header = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    log.info(f'Detected columns ({len(header)}): {header[:15]}…')

    def find_col(candidates):
        low_hdr = [h.lower() for h in header]
        for name in candidates:
            for i, h in enumerate(low_hdr):
                if name.lower() in h:
                    return i
        return None

    col_dist_id     = find_col(['district id', 'district-id', 'district number', 'cdn', 'district'])
    col_dist_name   = find_col(['district name', 'district-name', 'name'])
    col_year        = find_col(['school year', 'year', 'fiscal'])
    col_function    = find_col(['function-code', 'function code', 'function'])
    col_amount      = find_col(['actual-amount', 'actual amount', 'amount', 'total'])

    log.info(f'Column mapping: dist_id={col_dist_id}({header[col_dist_id] if col_dist_id is not None else "?"}), dist_name={col_dist_name}({header[col_dist_name] if col_dist_name is not None else "?"}), year={col_year}({header[col_year] if col_year is not None else "?"}), function={col_function}({header[col_function] if col_function is not None else "?"}), amount={col_amount}({header[col_amount] if col_amount is not None else "?"})')

    if None in (col_dist_id, col_dist_name, col_year, col_function, col_amount):
        log.error('Could not locate all required columns')
        log.error(f'Available columns: {header}')
        return None

    districts = defaultdict(lambda: {'name': '', 'years': defaultdict(lambda: defaultdict(float))})
    row_count = 0
    matched_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        if row_count % 100000 == 0:
            log.info(f'  Processed {row_count:,} rows ({matched_rows:,} facility-relevant)…')

        try:
            dist_id_raw = row[col_dist_id]
            dist_name   = str(row[col_dist_name]).strip() if row[col_dist_name] else ''
            year_raw    = row[col_year]
            func_raw    = row[col_function]
            amount_raw  = row[col_amount]

            if dist_id_raw is None or year_raw is None or func_raw is None or amount_raw is None:
                continue

            dist_id  = str(dist_id_raw).strip()
            year     = str(year_raw).strip()
            function = str(func_raw).strip().zfill(2)
            amount   = float(amount_raw)
        except (ValueError, TypeError, IndexError):
            continue

        if not dist_id or not year or function not in FACILITY_FUNCTIONS:
            continue

        matched_rows += 1
        # Normalize year format to something like "2024-2025"
        year_clean = year[:9] if len(year) >= 9 else year

        d = districts[dist_id]
        if not d['name'] and dist_name:
            d['name'] = dist_name
        d['years'][year_clean][FACILITY_FUNCTIONS[function]] += amount

    log.info(f'Parsed {row_count:,} rows, {matched_rows:,} were facility-relevant')
    log.info(f'Aggregated into {len(districts)} unique districts')
    return dict(districts)


# ── Build output ────────────────────────────────────────────────────────────
def build_output(districts_data):
    if not districts_data:
        return None

    all_years = set()
    for d in districts_data.values():
        all_years.update(d['years'].keys())
    sorted_years = sorted(all_years, reverse=True)[:YEARS_TO_KEEP]
    log.info(f'Keeping years: {sorted_years}')

    records = []
    for dist_id, d in districts_data.items():
        if not d['name']:
            continue

        history = []
        for y in sorted(sorted_years):
            year_data = d['years'].get(y, {})
            if not year_data:
                continue
            history.append({
                'year': y,
                'plant_maintenance':       round(year_data.get('plant_maintenance', 0), 0),
                'security_monitoring':     round(year_data.get('security_monitoring', 0), 0),
                'facilities_construction': round(year_data.get('facilities_construction', 0), 0),
            })

        if not history:
            continue

        latest = history[-1]
        total_facility = (
            latest['plant_maintenance'] +
            latest['security_monitoring'] +
            latest['facilities_construction']
        )
        if total_facility < 100000:
            continue

        records.append({
            'district_id':   dist_id,
            'district_name': d['name'],
            'latest_year':   latest['year'],
            'spending':      {**latest, 'total_facility': total_facility},
            'history':       history,
        })

    records.sort(key=lambda r: r['spending']['total_facility'], reverse=True)
    for i, r in enumerate(records):
        r['rank'] = i + 1

    return {
        'generated_at':   datetime.utcnow().isoformat() + 'Z',
        'source':         'Texas Education Agency PEIMS Summarized Actual Financial Data',
        'source_url':     TEA_DOWNLOADS_PAGE,
        'years_included': sorted_years,
        'notes': (
            'Annual actual expenditures per Texas ISD. Data is 1-2 years behind current. '
            'Function 51 = Plant M&O. Function 52 = Security. Function 81 = Facilities Construction.'
        ),
        'district_count': len(records),
        'districts':      records,
    }


# ── Main ────────────────────────────────────────────────────────────────────
def run():
    log.info('=' * 60)
    log.info('TEA District Finance Scraper')
    log.info('=' * 60)

    url = find_xlsx_url()
    if not url:
        return 1

    raw = download_tea_data(url)
    if not raw:
        return 1

    districts = parse_tea_xlsx(raw)
    if not districts:
        return 1

    output = build_output(districts)
    if not output:
        return 1

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, 'w') as f:
        json.dump(output, f, indent=2)

    log.info(f'✅ Wrote {output["district_count"]} districts to {OUTPUT_FILE}')
    log.info('Top 5 by facility spend (latest year):')
    for r in output['districts'][:5]:
        log.info(f'  #{r["rank"]}  {r["district_name"]:40s}  ${r["spending"]["total_facility"]/1e6:,.1f}M')

    return 0


if __name__ == '__main__':
    sys.exit(run())
